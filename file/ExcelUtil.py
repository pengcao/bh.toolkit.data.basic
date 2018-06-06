#!/usr/bin/env python
# encoding: utf-8
'''
@author: caopeng
@license: (C) Copyright 2017-2020, deamoncao@163.com 
@contact: deamoncao100@gmail.com
@software: garner
@file: ExcelUtil.py
@time: 2018/5/20 12:55
@desc:
'''

import xlrd
import xlwt
import openpyxl
import sys
from log.Logger import Logger

class ExcelUtil(object):

    __logger = Logger(sys.modules['__main__'])

    def readXlsxExcelSheetContentList(self,filePath,sheetName,head=True):
        '''
        读取excel里面某一个sheet name 的内容,返回的内容为一个List[list,list...]对象
        这个只支持.xlsx格式
        :param filePath:
        :param sheetName:
        :return:
        '''
        # sheet content list
        valueList = []
        try:
            wb = openpyxl.load_workbook(filePath)
            sheet = wb.get_sheet_by_name(sheetName)
            rowObjArray = list(sheet.rows)
            totalRow = sheet.max_row
            startRow = 0
            if head:
                startRow = 0
            else:
                startRow = 1
            for rowIndex in range(startRow,totalRow):
                row = rowObjArray[rowIndex]
                tmp = []
                for cell in range(len(row)):
                    tmp.append(row[cell].value)
                valueList.append(tmp)
        except Exception as ex:
            self.__logger.error("读取Excel文件失败！")
        return valueList

    def readXlsxExcelSheetContentDictList(self, filePath, sheetName):
        '''
        读取excel里面某一个sheet name 的内容,返回的内容为一个List[dict,dict...]对象
        这个只支持.xlsx格式
        :param filePath:
        :param sheetName:
        :return:
        '''
        # sheet content list
        valueList = []
        try:
            wb = openpyxl.load_workbook(filePath)
            sheet = wb.get_sheet_by_name(sheetName)
            # 生成器类型，不能使用索引
            rowObjArray = list(sheet.rows)
            totalRow = sheet.max_row
            startRow = 1
            headRow = rowObjArray[0]
            totalColumns = sheet.max_column
            # 读取
            for rowIndex in range(startRow, totalRow):
                row = rowObjArray[rowIndex]
                tmp = {}
                for cell in range(totalColumns):
                    tmp[headRow[cell].value] = row[cell].value
                valueList.append(tmp)
        except Exception as ex:
            self.__logger.error("读取Excel文件失败！")
        return valueList

    def readExcelSheetContentList(self,filePath,sheetName,head=True):
        '''
        读取excel里面某一个sheet name 的内容,返回的内容为一个List[list,list...]对象
        支持所有excel格式
        :param filePath:
        :param sheetName:
        :return:
        '''
        # sheet content list
        valueList = []
        try:
            wb = xlrd.open_workbook(filePath)
            sheet = wb.sheet_by_name(sheetName)
            # sheet.nrows
            totalRow = sheet.nrows
            startRow = 0
            if head:
                startRow = 0
            else:
                startRow = 1
            for rowIndex in range(startRow,totalRow):
                valueList.append(sheet.row_values(rowIndex))
        except Exception as ex:
            self.__logger.error("读取Excel文件失败！")
        return valueList

    def readExcelSheetContentDictList(self, filePath, sheetName):
        '''
        读取excel里面某一个sheet name 的内容,返回的内容为一个List[dict,dict...]对象
        支持所有excel格式
        :param filePath:
        :param sheetName:
        :return:
        '''
        # sheet content list
        valueList = []
        try:
            wb = xlrd.open_workbook(filePath)
            sheet = wb.sheet_by_name(sheetName)
            totalRow = sheet.nrows
            headRow = sheet.row_values(0)
            totalColumns = sheet.ncols
            # 读取
            for rowIndex in range(1, totalRow):
                row = sheet.row_values(rowIndex)
                tmp = {}
                for cell in range(totalColumns):
                    tmp[headRow[cell]] = row[cell]
                valueList.append(tmp)
        except Exception as ex:
            self.__logger.error("读取Excel文件失败！")
        return valueList

    def writeList2XlsxExcel(self,filePath,sheetName,valueList):
        '''
        将内容写入到Excel文件中，只支持xlsx格式的数据
        :param filePath:
        :param sheetName:
        :param valueList: list[list,list,...]
        :return:
        '''
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = sheetName
            for ii in range(0,len(valueList)):
                for jj in range(0,len(valueList[ii])):
                    sheet.cell(row = ii + 1,column = jj + 1, value = str(valueList[ii][jj]))
            wb.save(filePath)
        except Exception as ex:
            self.__logger.error("保存文件失败")

    def writeList2XlsxExcelWithHead(self,filePath,sheetName,valueList,headList):
        '''
        将内容写入到Excel文件中 只支持xlsx格式
        :param filePath:
        :param sheetName:
        :param valueList:
        :return:
        '''
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = sheetName
            # write head
            for jj in range(0, len(headList)):
                sheet.cell(row=1, column=jj + 1, value=str(headList[jj]))
            # write contents
            for ii in range(0,len(valueList)):
                for jj in range(0,len(valueList[ii])):
                    sheet.cell(row = ii + 2,column = jj + 1, value = str(valueList[ii][jj]))
            wb.save(filePath)
        except Exception as ex:
            self.__logger.error("保存文件失败")

    def writeDictList2XlsxExcelWithHead(self,filePath,sheetName,dictValueList,headList):
        '''
        将内容写入到Excel文件中 只支持xlsx格式
        :param filePath:
        :param sheetName:
        :param valueList:
        :return:
        '''
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = sheetName
            # write head
            totalColumns = len(headList)
            for jj in range(0, totalColumns):
                sheet.cell(row=1, column=jj + 1, value=str(headList[jj]))
            # write contents
            for ii in range(0,len(dictValueList)):
                dictRow = dictValueList[ii]
                for jj in range(0,totalColumns):
                    sheet.cell(row = ii + 2,column = jj + 1, value =dictRow[headList[jj]])
            wb.save(filePath)
        except Exception as ex:
            self.__logger.error("保存文件失败")

    def writeList2Excel(self,filePath,sheetName,valueList):
        '''
        将内容写入到Excel文件中,支持Excel所有格式
        :param filePath:
        :param sheetName:
        :param valueList:list[list,list,...]
        :return:
        '''
        try:
            wb = xlwt.Workbook()
            #
            sheet = wb.add_sheet(sheetName,cell_overwrite_ok=True)
            for ii in range(len(valueList)):
                for jj in range(len(valueList[ii])):
                    sheet.write(ii,jj,valueList[ii][jj])
            wb.save(filePath)
        except Exception as ex:
            self.__logger.error("保存文件失败")

    def writeList2ExcelWithHead(self,filePath,sheetName,valueList,headList):
        '''
        将内容写入到Excel文件中
        :param filePath:
        :param sheetName:
        :param valueList:
        :return:
        '''
        try:
            wb = xlwt.Workbook()
            #
            sheet = wb.add_sheet(sheetName,cell_overwrite_ok=True)
            for jj in range(len(headList)):
                sheet.write(0, jj, headList[jj])
            for ii in range(len(valueList)):
                for jj in range(len(valueList[ii])):
                    sheet.write(ii+1,jj,valueList[ii][jj])
            wb.save(filePath)
        except Exception as ex:
            self.__logger.error("保存文件失败")

    def writeDictList2ExcelWithHead(self, filePath, sheetName, valueDictList, headList):
        '''
        将内容写入到Excel文件中
        :param filePath:
        :param sheetName:
        :param valueDictList: dict list===> list[dict,dict,...]
        :param headList: head list
        :return:
        '''
        try:
            wb = xlwt.Workbook()
            #
            sheet = wb.add_sheet(sheetName, cell_overwrite_ok=True)
            totalColumns = len(headList)
            for jj in range(totalColumns):
                sheet.write(0, jj, headList[jj])
            for ii in range(len(valueDictList)):
                rowDict = valueDictList[ii]
                for jj in range(totalColumns):
                    sheet.write(ii + 1, jj, rowDict[headList[jj]])
            wb.save(filePath)
        except Exception as ex:
            self.__logger.error("保存文件失败")

    def combine2ExcelContent(self,file01Path,sheet01Name,excel01Index,file02Path,sheet02Name,excel02Index,
                             resultFilePath,resultSheetName):
        '''
         将两个Excel中的内容按照某一列进行join操作（实现类似于SQL的joinc操作）
        :param file01Path:
        :param sheet01Name:
        :param excel01Index:
        :param file02Path:
        :param sheet02Name:
        :param excel02Index:
        :param resultFilePath:
        :param resultSheetName:
        :return:
        '''
        excel01ValueList = self.readExcelSheetContentList(file01Path,sheet01Name)
        excel02ValueList = self.readExcelSheetContentList(file02Path,sheet02Name)

        resultValueList = []
        for ii in range(len(excel01ValueList)):
            tmp = []
            for jj in range(len(excel02ValueList)):
                if str(excel01ValueList[ii][excel01Index])==str(excel02ValueList[jj][excel02Index]):
                    for xx in range(len(excel01ValueList[ii])):
                        if xx == excel01Index:
                            continue
                        excel02ValueList[jj].append(excel01ValueList[ii][xx])
                    tmp = excel02ValueList[jj]
                    resultValueList.append(tmp)
                    break
        self.write2Excel(resultFilePath,resultSheetName,resultValueList)


if __name__ == '__main__':
    print('=== start......')
    excelUtil = ExcelUtil()
    sheetName = 'Sheet1'
    excelPath = 'C:/develop/test0001.xlsx'
    rntnDictList = excelUtil.readXlsxExcelSheetContentDictList(excelPath,sheetName)
    print('=== dict excel ',rntnDictList)
    # write
    headList = ['id','name','age']
    svExcel11Path = 'C:/develop/test_xsls_3001.xlsx'
    excelUtil.writeDictList2XlsxExcelWithHead(svExcel11Path,sheetName,rntnDictList,headList)

